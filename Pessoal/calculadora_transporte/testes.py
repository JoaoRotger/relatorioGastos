import pandas as pd
import contaDias as cd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter  # Importe esta linha

# Obtém os dias do mês
data = cd.getDate()
mes = cd.getMonthStr()
dias_do_mes = cd.dias_do_mes()

# Criação do DataFrame
tabela = pd.DataFrame({'Data': dias_do_mes, 'Motivo': 'Transporte Ida e Volta', 'Valor': 8.80})
tabela = tabela[['Data', 'Motivo', 'Valor']]
tabela = tabela.reset_index(drop=True)

# Salva o DataFrame em uma planilha Excel
nome_arquivo = f'Relatorio_de_Gastos_{data[1]}{data[2]}.xlsx'
tabela.to_excel(nome_arquivo, sheet_name=f'{data[1]}{data[2]}', index=False)

# Carrega a planilha Excel
workbook = load_workbook(nome_arquivo)
sheet = workbook[f'{mes}{data[2]}']

# Ajusta a largura das colunas com base no conteúdo
for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

# Salva as alterações na planilha Excel
workbook.save(nome_arquivo)

print(f'Planilha "{nome_arquivo}" salva com sucesso!')
