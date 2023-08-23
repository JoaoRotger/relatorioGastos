import pandas as pd
import contaDias as cd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
import json

confsPath = os.path.join(os.path.dirname(__file__), "conf.json") 
with open('conf.json', 'r') as file:
  config = json.load(file)
  valorPassagem = config["valorPassagem"]

data = cd.getDate()
dias_do_mes = cd.dias_do_mes()
tabela = pd.DataFrame({'Data': dias_do_mes, 'Motivo': 'Transporte Ida e Volta', 'Valor': 8.80})
tabela = tabela[['Data', 'Motivo', 'Valor']]
tabela = tabela.reset_index(drop=True)

valor_total = cd.quantDias(data[1], data[2]) * valorPassagem
linha_total = pd.DataFrame({'Data': 'Valor Total', 'Motivo': '', 'Valor': valor_total}, index=[-1])
tabela = pd.concat([tabela, linha_total]).reset_index(drop=True)

nome_arquivo = f'Relatorio_de_Gastos_{cd.getMonthStr()}{data[2]}.xlsx'
tabela.to_excel(nome_arquivo, sheet_name=f'{data[1]}{data[2]}', index=False)

workbook = load_workbook(nome_arquivo)
sheet = workbook[f'{data[1]}{data[2]}']

for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

font_negrito = Font(bold=True)
sheet[f'C{len(tabela)+2}'].font = font_negrito
sheet[f'D{len(tabela)+2}'].font = font_negrito

workbook.save(nome_arquivo)

print(f'Planilha "{nome_arquivo}" salva com sucesso!')
